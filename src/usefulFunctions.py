from openpyxl import load_workbook
from pathlib import Path
import sys
import datetime
import os
import shutil

def isNaN(num):
    return num!= num

def getCurrentPath():   
    config_name = 'myapp.cfg'
    # determine if application is a script file or frozen exe
    if getattr(sys, 'frozen', False):
        application_path = os.path.dirname(sys.executable)
    elif __file__:
        application_path = os.path.dirname(__file__)
    application_path2 = Path(application_path)
    return application_path2.absolute()

currentPathFolder = getCurrentPath()
currentPathParentFolder = Path(getCurrentPath()).parent
currentPathGrandpaFolder = Path(currentPathParentFolder).parent
currentPathGrandpaFolderParent = Path(currentPathGrandpaFolder).parent
logPath = os.path.join(currentPathGrandpaFolder,"log.txt")

def today():
    fullTime = datetime.datetime.now(datetime.timezone.utc) + datetime.timedelta(hours=-4)
    currenteDateStr = fullTime.strftime("%d.%m.%Y")
    return currenteDateStr

def today2():
    fullTime = datetime.datetime.now(datetime.timezone.utc) + datetime.timedelta(hours=-4)
    currenteDateStr = fullTime.strftime("%d%m%Y")
    return currenteDateStr

def today3():
    fullTime = datetime.datetime.now(datetime.timezone.utc) + datetime.timedelta(hours=-4)
    currenteDateStr = fullTime.strftime("%d.%m.%Y-%H %M %S")
    return currenteDateStr




def xlsxFormatting(x, n):
    z1 = ''.join([today(), '-F'])
    wb1 = load_workbook(x)
    xlsxFormatedFolder = os.path.join(currentPathParentFolder,"Cuentas recaudadoras 2")
    xlsxFormatedFolder = os.path.join(xlsxFormatedFolder,z1)
    try:        
        os.mkdir(xlsxFormatedFolder)
    except Exception as e:
        print('El archivo ya ha sido creado')
    
    match n:
        case 1:
            xlsxFormatedPath = os.path.join(xlsxFormatedFolder,"CUENTA ETV-F.xlsx")   

        case 2:
             xlsxFormatedPath = os.path.join(xlsxFormatedFolder,"CUENTA BANCO-F.xlsx")

    wb1.save(xlsxFormatedPath)
    wb2 = load_workbook(xlsxFormatedPath)
    ws2 = wb2['CAJAS RECAUDADORAS']
    mergeRangesList = []
    for i in ws2.merged_cell_ranges:
        i = str(i)
        j = i.index(':')
        #print(i, ' ', j)
        if i[0]==i[j+1] and i[0]=='D' or i[0]==i[j+1]=='E' and i[0]=='E':
            if int(i[1:j])>=3:
                mergeRangesList.append(i)
    #print(mergeRangesList[:])

    for k in mergeRangesList:
        #print(k)
        ws2.unmerge_cells(k)
        l = k.index(':')
        
        a = k[1:l]
        b = k[l+2:]
        #print(a)
        #print(b)
        a = int(a)
        b = int(b)
        a+=1
        b+=1
        
        for m in range(a, b):
            n = ''.join([k[0], str(m)])
            #print( ws2[f'{n}'])
            #print(ws2[k[:l]])
            ws2[f'{n}'] = ws2[k[:l]].value



    wb2.save(xlsxFormatedPath)
    return xlsxFormatedPath

def report(n, asignacion1, accountNumberStr2, accountNumberStr1):
    global errorList
    errorList = []

    match n:
        case 1:
            p = asignacion1 + '-' + accountNumberStr2 + '-' + accountNumberStr1 + ' fue migrado correctamente'
            errorList.append(p)
            # print(p)
            writeLog('\n', p, logPath)
        case 2:
            p = asignacion1 + '-' + accountNumberStr2 + '-' + accountNumberStr1 + ' no fue migrado correctamente, revisar manualmente'
            errorList.append(p)
            # print(p)
            writeLog('\n', p, logPath)
        case _:
            p = 'Error-ingresó un número incorrecto a la función report'
            errorList.append(p)
            # print(p)
            writeLog('\n', p, logPath)

def writeLog(s,log,rut):
    txtfolder=os.path.dirname(rut)
    pathLog=os.path.join(txtfolder, "logs.txt")
    line=s+str(log)
    print(log)
    with open(pathLog, 'a') as file:
        file.write(line)

def fecha_a_dia(fecha):
    dias_de_la_semana = ["Lunes","Martes","Miercoles","Jueves","Viernes","Sabado","Domingo"]
    fecha_split = fecha.split(".")
    fecha_formateada = datetime.date(int(fecha_split[2]), int(fecha_split[1]), int(fecha_split[0]))
    dia = dias_de_la_semana[fecha_formateada.weekday()]
    return dia

def copyANDeraseFile(fileName):
    todayPath = os.path.join(currentPathParentFolder,"Cuentas recaudadoras")
    fileFromTodayPathFrom = os.path.join(todayPath,fileName)
    todayPath = os.path.join(todayPath,today())
    if not os.path.exists(todayPath):
        os.mkdir(todayPath)
    fileName = today3() + ' ' + fileName
    fileFromTodayPathTo = os.path.join(todayPath,fileName)
    shutil.copyfile(fileFromTodayPathFrom, fileFromTodayPathTo)
    os.remove(fileFromTodayPathFrom)

def copyFile(fileName):
    fileNamePathFrom = os.path.join(currentPathParentFolder, fileName)
    fileNamePathTo = os.path.join(currentPathParentFolder, "Cuentas recaudadoras")
    fileNamePathTo = os.path.join(fileNamePathTo, fileName)
    shutil.copyfile(fileNamePathFrom, fileNamePathTo)

def ndocTOxlsx(asignacionNdocMigrated, rec, xlsx, logPath):
    xlsx = xlsx + '.xlsx'
    xlsxPath = os.path.join(currentPathParentFolder, 'Migraciones', xlsx)
    try:
        wb = load_workbook(xlsxPath)
    except:
        writeLog('\n', f'El archivo {xlsx} no existe en el directorio o está abierto.', logPath)
    try:
        ws = wb[rec]
        for l in asignacionNdocMigrated:
            counter = 0
            for i in range(1, ws.max_row+1):
                asignacion = l[0]
                ndoc = l[1]
                if asignacion == ws.cell(row=i, column=1).value:
                    ws[f'K{i}'] = ndoc
                    counter = 1
                    break      
            if counter == 0:
                writeLog('\n', f'La asignación {asignacion} no existe en la hoja {rec} del archivo {xlsx}.', logPath)   
        wb.save(xlsxPath)
        wb.close()
    except:
        writeLog('\n', f'La hoja {rec} no existe en el archivo {xlsx}.', logPath)

def add0(num):
  num_str = str(num)
  if len(num_str) == 1:
    return '0' + num_str
  return num_str










if __name__=='__main__': 
  ndocTOxlsx('CCAJ-LP08/298/22', '13161561565', 'AG. ACHUMANI', 'MIGRACIONES SGV DICIEMBRE 2022 28.12.2022', r'C:\Users\crist\OneDrive - UNIVERSIDAD NACIONAL DE INGENIERIA\Venado\Cris\Bot5\Cuentas recaudadoras\logs.txt')