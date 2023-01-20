import win32com.client
import subprocess
import time
from datetime import datetime, timedelta
from openpyxl import load_workbook
import re
import os
from usefulFunctions import currentPathParentFolder, currentPathGrandpaFolder, today, writeLog, fecha_a_dia, copyANDeraseFile, copyFile, ndocTOxlsx, add0
import pandas as pd
import gc

#poo desde video 26

class sapInterfaceJob():
    def __init__(self):
        self.paths = {}
        self.login = {}
    
        self.sapGuiAuto = None
        self.application = None
        self.connection = None
        self.session = None
        self.paths = None
        self.dailyMigrationAccountsPath = None
        self.login = None
        self.proc = None
        self.r = None
        self.rec = None
        self.txtCabDoc = None
        self.importe = None
        self.asignacion = None
        self.fullAsignacion = None
        self.dist = None
        self.texto = None
        self.moneda = None
        self.a = None
        self.f = None
        self.c = None
        self.im = None
        self.fecha = None
        self.fecha2 = None
        self.ct = None
        self.check = None
        self.wb2 = None
        self.wsDist = None
        self.wsAg = None
        self.wb3 = None
        self.wsNames = None

        self.ws2 = None

        self.accountNumber1 = None
        self.accountNumberStr1 = None
        self.accountNumber2 = None
        self.accountNumberStr2 = None
        self.bank = None
        self.layout = None
        self.asignaciones = []
        self.ndocs = []
        self.fechas = []
        self.cts = []
        self.importes = []        
        self.textos = []
        self.checks = []
        self.fullAsignaciones = []
        self.dists = []
        self.finalTexts = []
        self.tMigracion = None
        self.xlsxRange = None

        self.wholeParametersList = []
        self.approvedAssignments = []
        self.approvedNdocs = []
        self.approvedFechas = []
        self.approvedCts = []
        self.approvedImportes = []
        self.approvedTextos = []
        self.approvedChecks = []
        self.approvedFullAsignaciones = []
        self.approvedDists = []
        self.approvedFinalTexts = []
        self.docf = None

        self.approvedParametersList = []

        self.listOfNames = []
        self.listOfImporteIndex = []
        self.listOfFechaIndex = []


        self.currentPathParentFolder = currentPathParentFolder
        self.currentPathGrandpaFolder = currentPathGrandpaFolder
        self.logPath = os.path.join(self.currentPathParentFolder,"Cuentas recaudadoras", "log.txt")
        # self.currentPathGrandpaFolder = 'C:\\Users\\crist\\OneDrive - UNIVERSIDAD NACIONAL DE INGENIERIA\\Venado\\Cris\\Traslado de tesoreria B5'
        self.directo = None
        self.ETVflow = None
        self.xlsxMigracion = None
        
        self.changeThePeriod = False
        self.i_0 = 3
        self.j_0 = 0
        self.i = 3
        self.j = 0
        self.jMax = 3
        self.k = 7
        self.rowCount = 0

        self.exec = None
        

    def startSAP(self):
        
        configXlsx=os.path.join(self.currentPathParentFolder,"config.xlsx")
        wb = load_workbook(configXlsx)
        ws = wb['Rutas']
        wsConfig = wb['parametrosInicio']

        self.paths = {'SAPPath': ws['B2'].value}
               
        
        self.login = {'user': wsConfig['B1'].value,
                 'psw': wsConfig['B2'].value,
                 'environment': wsConfig['B3'].value,
                 'fecha': wsConfig['B5'].value,
                 'periodo': wsConfig['B6'].value,
                 'layout': wsConfig['B8'].value,
                 'xlsx migracion': wsConfig['B10'].value}
        
        wb.close()
                
        self.layout = self.login['layout']
        self.layout = self.layout.replace(" ","")
        self.xlsxMigracion = self.login['xlsx migracion']

        if self.login['fecha'] == None:
            self.login['fecha'] = today()
        else:
            self.changeThePeriod = True
      
        self.proc = subprocess.Popen([self.paths['SAPPath'], '-new-tab'])
        time.sleep(2)
        try: 
            self.sapGuiAuto = win32com.client.GetObject('SAPGUI')
        except:
            self.proc.kill()
            time.sleep(2)
            self.proc = subprocess.Popen([self.paths['SAPPath'], '-new-tab'])
            time.sleep(2)
            self.sapGuiAuto = win32com.client.GetObject('SAPGUI')

        self.application = self.sapGuiAuto.GetScriptingEngine
        self.connection = self.application.OpenConnection(self.login['environment'], True)
        self.session = self.connection.Children(0)

        self.session.findById("wnd[0]/usr/txtRSYST-BNAME").text = self.login['user']
        self.session.findById("wnd[0]/usr/pwdRSYST-BCODE").text = self.login['psw']
        self.session.findById("wnd[0]").sendVKey(0)

    def getFbl3nMenu(self):
        self.session.EndTransaction()
        self.session.findById("wnd[0]/tbar[0]/okcd").text = "fbl3n"
        self.session.findById("wnd[0]").sendVKey(0)

    def chargeXlsxSheet(self):
        self.dailyMigrationAccountsPath=os.path.join(self.currentPathParentFolder,"Cuentas Recaudadoras")
        self.dailyMigrationAccountsPath=os.path.join(self.dailyMigrationAccountsPath,"CUENTAS DE CAJA IVSA.xlsx") 
        #EL self.wb ES USADO PARA LEER EL CONFIG.
        self.wb2 = load_workbook(self.dailyMigrationAccountsPath)
        self.wsDist = self.wb2['DISTRIBUIDORAS']
        self.wsAg = self.wb2['AGENCIAS']
        

    def chargeListOfNames(self):
        y = os.path.join(self.currentPathParentFolder,"BASE DE DATOS DIST.xlsx")
        self.wb3 = load_workbook(y)
        self.wsNames = self.wb3['Hoja1']
        for i in range(2, 68):
            cellName = self.wsNames[f'D{i}'].value
            cellName = cellName.replace(" ", "")
            if cellName != None and cellName != "":
                self.listOfNames.append(cellName)

        self.wb3.close()
        

    def getExcelRange(self):
        xlsxCellsRange = []

        while True:
            self.accountNumber1 = self.ws2[f'C{self.i}'].value
            self.accountNumberStr1 = str(self.accountNumber1).replace(' ', '')
            self.accountNumber2 = self.ws2[f'D{self.i}'].value
            self.accountNumberStr2 = str(self.accountNumber2).replace(' ', '')

            if len(self.accountNumberStr1)==9 and len(self.accountNumberStr2)==9 and type(self.accountNumber1)== int and type(self.accountNumber2)== int:
                xlsxCellsRange.append(self.i)
                self.i+=1
            else:
                self.i+=1
                self.j+=1
                if self.j > self.jMax:
                    break
                else:
                    continue
        return xlsxCellsRange
            

    def getWholeParametersList(self):
        self.wholeParametersList = []
        self.rowCount = self.session.findById('wnd[0]/usr/cntlGRID1/shellcont/shell/shellcont[1]/shell').RowCount
        #self.rowCount-=3
        self.rowCount = min([self.rowCount, 62])

        for k in range(self.rowCount):
            self.k = k
            self.getRowInformation(self.k)
            if self.check == 0:
                self.asignaciones.append(self.asignacion)
                self.ndocs.append(self.ndoc)
                self.fechas.append(self.fecha)
                self.cts.append(self.ct)
                self.importes.append(self.importe)
                self.textos.append(self.texto1)
                self.checks.append(self.check)
                self.fullAsignaciones.append(self.fullAsignacion)
                self.dists.append(self.dist)
                self.finalTexts.append(self.texto)
           
        self.wholeParametersList.append(self.asignaciones)
        self.wholeParametersList.append(self.ndocs)
        self.wholeParametersList.append(self.fechas)
        self.wholeParametersList.append(self.cts)
        self.wholeParametersList.append(self.importes)
        self.wholeParametersList.append(self.textos)
        self.wholeParametersList.append(self.checks)
        self.wholeParametersList.append(self.fullAsignaciones)
        self.wholeParametersList.append(self.dists)
        self.wholeParametersList.append(self.finalTexts)

        self.asignaciones = []
        self.ndocs = []
        self.fechas = []
        self.cts = []
        self.importes = []
        self.textos = []
        self.checks = []
        self.fullAsignaciones = []
        self.dists = []
        self.finalTexts = []

        return self.wholeParametersList

    def IndexOfRepitedImport(self, importe, list2):
        
        for i, element in enumerate(list2):
            if importe == element.replace('-',''):
                self.listOfImporteIndex.append(i)
        return self.listOfImporteIndex

    def IndexOfRepitedFecha(self, fecha, list2):

        if fecha_a_dia(fecha) == 'Sabado':
            fecha = datetime.strptime(fecha, '%d.%m.%Y')
            fecha+=timedelta(days = 2)
            fecha = f"{add0(fecha.day)}.{add0(fecha.month)}.{add0(fecha.year)}"

        else:
            fecha = datetime.strptime(fecha, '%d.%m.%Y')
            fecha+=timedelta(days = 1)
            fecha = f"{add0(fecha.day)}.{add0(fecha.month)}.{add0(fecha.year)}"

        for i, element in enumerate(list2):
            if fecha == element:
                self.listOfFechaIndex.append(i)

        return self.listOfFechaIndex

    def commons(self, list1, list2):
        return list(set(list1).intersection(list2))
                      

    def lastValidationChecker(self, preApprovedParametersList, parametersList2):
        approvedParametersList = []
        fechaIndexs = []
        importeIndexs = []
        asignaciones = []
        ndocs = []
        dates = []
        cts = []
        imports = []
        texts = []
        checks = []
        fullAsignaciones = []
        dists = []
        finalTexts = []

        fechas = preApprovedParametersList[2]
        importes = preApprovedParametersList[4]
        fechas2 = parametersList2[2]
        importes2 = parametersList2[4]
        textos2 = parametersList2[5]
        approvedIndexs = []

        for fecha in fechas:
            i = fechas.index(fecha)
            importe = importes[i]
            fechaIndexs = self.IndexOfRepitedFecha(fecha, fechas2)
            importeIndexs = self.IndexOfRepitedImport(importe, importes2)

        commonIndexs = self.commons(fechaIndexs, importeIndexs)
        

        for j in commonIndexs:

            if ':' in textos2[j]:
                n = textos2[j].index(':')
                texto = textos2[j][n+1:]

            else:
                texto = textos2[j]
                x = re.findall(r'\d+-*\d+\w*\s', texto)

                for i1 in x:
                    texto = texto.replace(i1, '')

                x2 = re.findall(r'[^a-zA-Z0-9\s]', texto)

                for i2 in x2:
                    texto = texto.replace(i2, '')
                
                

            if '(' in texto:
                m = texto.index('(')
                texto = texto[:m]
            texto = texto.strip()

            splitList = re.split(r'\s', texto)
            splitList = splitList[:3]
            
            for name in self.listOfNames:
                if splitList[0] in name and splitList[1] in name and splitList[2] in name:
                    i = importes.index(importes2[j].replace('-',''))
                    approvedIndexs.append(i)
                    break                

        for k in approvedIndexs:
            asignaciones.append(preApprovedParametersList[0][k])
            ndocs.append(preApprovedParametersList[1][k])
            dates.append(preApprovedParametersList[2][k])
            cts.append(preApprovedParametersList[3][k])
            imports.append(preApprovedParametersList[4][k])
            texts.append(preApprovedParametersList[5][k])
            checks.append(preApprovedParametersList[6][k])
            fullAsignaciones.append(preApprovedParametersList[7][k])
            dists.append(preApprovedParametersList[8][k])
            finalTexts.append(preApprovedParametersList[9][k])

        approvedParametersList.append(asignaciones)
        approvedParametersList.append(ndocs)
        approvedParametersList.append(dates)
        approvedParametersList.append(cts)
        approvedParametersList.append(imports)
        approvedParametersList.append(texts)
        approvedParametersList.append(checks)
        approvedParametersList.append(fullAsignaciones)
        approvedParametersList.append(dists)
        approvedParametersList.append(finalTexts)

        return approvedParametersList   


    def lastValidationChecker2(self, preApprovedParametersList, parametersList2):
        approvedParametersList = []
        asignaciones = []
        ndocs = []
        dates = []
        cts = []
        imports = []
        texts = []
        checks = []
        fullAsignaciones = []
        dists = []
        finalTexts = []
        approvedIndexs = []
        for i, element in enumerate(preApprovedParametersList[0]):
            for j, element2 in enumerate(parametersList2[0]):
                fecha = preApprovedParametersList[2][i]
                if fecha_a_dia(fecha) == 'Viernes':
                    fecha = datetime.strptime(fecha, '%d.%m.%Y')
                    fecha+=timedelta(days = 3)
                    fecha = f"{fecha.day}.{fecha.month}.{fecha.year}"

                else:
                    fecha = datetime.strptime(fecha, '%d.%m.%Y')
                    fecha+=timedelta(days = 2)
                    fecha = f"{fecha.day}.{fecha.month}.{fecha.year}"

                if element == element2 and fecha == parametersList2[2][j] and preApprovedParametersList[4][i] == parametersList2[4][j].replace('-',''):
                    texto = parametersList2[5][j]
                    if ':' in texto:
                        n = texto.index(':')
                        texto = texto[n+1:]
                        texto = texto.replace(' ', '')
                    
                    else:
                        texto = texto[13:]
                        texto = texto.replace(' ', '')
                    
                    if '(' in texto:
                        m = texto.index('(')
                        texto = texto[:m]
                        texto = texto.replace(' ', '')

                    for name in self.listOfNames:
                        if texto in name:
                            approvedIndexs.append(i)

        for k in approvedIndexs:
            asignaciones.append(preApprovedParametersList[0][k])
            ndocs.append(preApprovedParametersList[1][k])
            dates.append(preApprovedParametersList[2][k])
            cts.append(preApprovedParametersList[3][k])
            imports.append(preApprovedParametersList[4][k])
            texts.append(preApprovedParametersList[5][k])
            checks.append(preApprovedParametersList[6][k])
            fullAsignaciones.append(preApprovedParametersList[7][k])
            dists.append(preApprovedParametersList[8][k])
            finalTexts.append(preApprovedParametersList[9][k])

        approvedParametersList.append(asignaciones)
        approvedParametersList.append(ndocs)
        approvedParametersList.append(dates)
        approvedParametersList.append(cts)
        approvedParametersList.append(imports)
        approvedParametersList.append(texts)
        approvedParametersList.append(checks)
        approvedParametersList.append(fullAsignaciones)
        approvedParametersList.append(dists)
        approvedParametersList.append(finalTexts)

        return approvedParametersList
                    



    def wichMigraVerification2(self, preApprovedParametersList):
        self.getFbl3nMenu()
        self.getAccountTable2()
        parametersList2 = self.getWholeParametersList()
        self.listOfFechaIndex = []
        self.listOfImporteIndex = []
        approvedParametersList = self.lastValidationChecker(preApprovedParametersList, parametersList2)

        return approvedParametersList


    def wichMigraVerification(self, wholeParametersList):
        for assigment in wholeParametersList[0]:
            if wholeParametersList[0].count(assigment) == 1:               
                n = wholeParametersList[0].index(assigment)
                ndoc = wholeParametersList[1][n]
                fecha = wholeParametersList[2][n]
                ct = wholeParametersList[3][n]
                importe = wholeParametersList[4][n]
                texto = wholeParametersList[5][n]
                check = wholeParametersList[6][n]
                fullAsignacion = wholeParametersList[7][n]
                dist = wholeParametersList[8][n]
                finalText = wholeParametersList[9][n]
                if ct == '40' and check == 0:
                    self.approvedAssignments.append(assigment)
                    self.approvedNdocs.append(ndoc)
                    self.approvedFechas.append(fecha)
                    self.approvedCts.append(ct)
                    self.approvedImportes.append(importe)
                    self.approvedTextos.append(texto)
                    self.approvedChecks.append(check)
                    self.approvedFullAsignaciones.append(fullAsignacion)
                    self.approvedDists.append(dist)
                    self.approvedFinalTexts.append(finalText)

        approvedParametersList = []
        approvedParametersList.append(self.approvedAssignments)
        approvedParametersList.append(self.approvedNdocs)
        approvedParametersList.append(self.approvedFechas)
        approvedParametersList.append(self.approvedCts)
        approvedParametersList.append(self.approvedImportes)
        approvedParametersList.append(self.approvedTextos)
        approvedParametersList.append(self.approvedChecks)
        approvedParametersList.append(self.approvedFullAsignaciones)
        approvedParametersList.append(self.approvedDists)
        approvedParametersList.append(self.approvedFinalTexts)
        self.approvedAssignments = []
        self.approvedNdocs = []
        self.approvedFechas = []
        self.approvedCts = []
        self.approvedImportes = []
        self.approvedTextos = []
        self.approvedChecks = []
        self.approvedFullAsignaciones = []
        self.approvedDists = []
        self.approvedFinalTexts = []
        
        return approvedParametersList

    def verificationBeforeAccountChange(self, nDocsMigrated, approvedParametersList, wholeparametersList):
        counter = 0
        for ndoc in nDocsMigrated:
            if ndoc in wholeparametersList[1]:
                n = wholeparametersList[1].index(ndoc)
                importe1 = wholeparametersList[4][n]
                importe1 = importe1.replace(' ', '')
                importe1 = importe1.replace('-', '')
                importe2 = approvedParametersList[4][counter]
                importe2 = importe2.replace(' ', '')
                importe2 = importe2.replace('-', '')
                if importe1 == importe2:
                    x = "%s La operación de asignación: %s fue migrada correctamente" %(today(), approvedParametersList[0][counter])
                    writeLog('\n', x, self.logPath)
                else:
                    y = "%s La operación de asignación: %s ERROR en importe migrado, revisar manualmente" %(today(), approvedParametersList[0][counter])
                    writeLog('\n', y, self.logPath)
            else:
                z = f'La operación de asignación: {approvedParametersList[0][counter]} FALLO en el guardado o pérdida de datos, revisar manualmente'
                y = "%s La operación de asignación: %s FALLO en el guardado o pérdida de datos, revisar manualmente" %(today(), approvedParametersList[0][counter])
                writeLog('\n', z, self.logPath)
            counter+=1    
             
# PROCESO -------------------------------------------------------------
    def migration(self, rowList):                
        self.session.EndTransaction()

        self.session.findById("wnd[0]/tbar[0]/okcd").text = "f-02"
        self.session.findById("wnd[0]").sendVKey(0)

        if self.changeThePeriod:
             self.session.findById("wnd[0]/usr/ctxtBKPF-BUDAT").text = self.login['fecha']
             self.session.findById("wnd[0]/usr/txtBKPF-MONAT").text = self.login['periodo']

        self.session.findById("wnd[0]/usr/ctxtBKPF-BLDAT").text = self.login['fecha']

        recaudadora = self.rec
        recaudadora = recaudadora.replace('CENTRAL', '')
        self.session.findById("wnd[0]/usr/txtBKPF-XBLNR").text = recaudadora
        try:
            self.session.findById("wnd[0]/usr/txtBKPF-BKTXT").text = self.txtCabDoc
        except:
            self.txtCabDoc = self.txtCabDoc.replace('TRASLADO', 'TRASL')
            self.session.findById("wnd[0]/usr/txtBKPF-BKTXT").text = self.txtCabDoc
        self.session.findById("wnd[0]/usr/ctxtRF05A-NEWKO").text = self.accountNumberStr2
        
        self.session.findById("wnd[0]/tbar[0]/btn[0]").press()
        
        try:
            self.session.findById("wnd[0]/usr/txtBSEG-WRBTR").text = rowList[4]
        except:
            periodFail = self.session.findById("wnd[0]/sbar/pane[0]").text
            self.session.endTransaction()

            raise Exception(periodFail)
        self.session.findById("wnd[0]/usr/txtBSEG-ZUONR").text = rowList[7]
        self.session.findById("wnd[0]/usr/ctxtBSEG-SGTXT").text = rowList[9]
        self.session.findById("wnd[0]/usr/ctxtRF05A-NEWBS").text = '50'
        self.session.findById("wnd[0]/usr/ctxtRF05A-NEWKO").text = self.accountNumberStr1
        self.session.findById("wnd[0]/tbar[0]/btn[0]").press()

        self.session.findById("wnd[0]/usr/txtBSEG-WRBTR").text = rowList[4]
        self.session.findById("wnd[0]/usr/txtBSEG-ZUONR").text = rowList[7]
        self.session.findById("wnd[0]/usr/ctxtBSEG-SGTXT").text = rowList[9]
        self.session.findById("wnd[0]/mbar/menu[0]/menu[3]").select()

        validacion = self.session.findById("wnd[0]/usr/txtRF05A-AZSAL").text
        validacion = str(validacion)
        validacion = validacion.replace(' ', '')
        validacion = validacion.replace('.', '')
        validacion = validacion.replace(',', '.')
        validacion = float(validacion)
        if validacion == 0:
            x = f'Validación de saldo 0 correcto en asignación: {rowList[0]}'
            writeLog('\n', x, self.logPath)
        else:
            y = f'ERROR DE VALIDACIÓN DE SALDO 0 EN ASIGNACIÓN: {rowList[0]}'
            writeLog('\n', y, self.logPath)       
   
        self.session.findById("wnd[0]/tbar[0]/btn[11]").press()

        self.docf = self.session.findById("wnd[0]/sbar/pane[0]").text
        self.docf = self.docf.replace(' ', '')
        self.docf = self.docf[4:13]
        if len(self.docf) != 9:
            self.docf = 'No hay N° doc.'

        self.session.EndTransaction()
                
    def getAccountTable(self):
        self.session.findById("wnd[0]/usr/ctxtSD_SAKNR-LOW").text = self.accountNumberStr1
        self.session.findById("wnd[0]/usr/ctxtSD_BUKRS-LOW").text = "GV01"
        if self.layout is not None and self.layout != '':
            self.session.findById("wnd[0]/usr/ctxtPA_VARI").text = self.layout
        self.session.findById("wnd[0]/usr/ctxtSD_BUKRS-LOW").setFocus
        self.session.findById("wnd[0]/tbar[1]/btn[8]").press()

    def getAccountTable2(self):
        self.session.findById("wnd[0]/usr/ctxtSD_SAKNR-LOW").text = self.accountNumberStr2
        self.session.findById("wnd[0]/usr/ctxtSD_BUKRS-LOW").text = "GV01"
        if self.layout is not None and self.layout != '':
            self.session.findById("wnd[0]/usr/ctxtPA_VARI").text = self.layout
        self.session.findById("wnd[0]/usr/ctxtSD_BUKRS-LOW").setFocus
        self.session.findById("wnd[0]/tbar[1]/btn[8]").press()
        
        

    def getRowInformation(self, k):
        self.asignacion = None
        self.ndoc = None
        self.fecha = None
        self.ct = None
        self.importe = None
        self.txt = None
        self.check = None

        self.fullAsignacion = None
        self.dist = None
        
        self.asignacion = self.session.findById('wnd[0]/usr/cntlGRID1/shellcont/shell/shellcont[1]/shell').GetCellValue(k, 'ZUONR')
        self.ndoc = self.session.findById('wnd[0]/usr/cntlGRID1/shellcont/shell/shellcont[1]/shell').GetCellValue(k, 'BELNR')
        self.fecha = self.session.findById('wnd[0]/usr/cntlGRID1/shellcont/shell/shellcont[1]/shell').GetCellValue(k, 'BLDAT')
        self.ct = self.session.findById('wnd[0]/usr/cntlGRID1/shellcont/shell/shellcont[1]/shell').GetCellValue(k, 'BSCHL')
        self.importe = self.session.findById('wnd[0]/usr/cntlGRID1/shellcont/shell/shellcont[1]/shell').GetCellValue(k, 'DMSHB')
        self.texto1 = self.session.findById('wnd[0]/usr/cntlGRID1/shellcont/shell/shellcont[1]/shell').GetCellValue(k, 'SGTXT')
        self.check = self.session.findById('wnd[0]/usr/cntlGRID1/shellcont/shell/shellcont[1]/shell').GetCellValue(k, 'ICO_AUGP')
        if 'Pendientes' in self.check:
            self.check = 0
        else:
            self.check = 1
        self.asignacion = str(self.asignacion).replace(' ', '')
        self.fullAsignacion = self.asignacion
        self.dist = self.asignacion[5:7]
        if self.asignacion.count('/') > 1:
            self.asignacion = self.asignacion[::-1]
            try:
                n = self.asignacion.index('/')
                self.asignacion = self.asignacion[n:]
            except:
                report = 'La asignación no tiene /'
            self.asignacion = self.asignacion[::-1]
        self.ndoc = str(self.ndoc).replace(' ', '')
        self.fecha = str(self.fecha).replace(' ', '')
        try:
            l = self.fecha.index('.')
            self.fecha2 = self.fecha[:l+3]
        except:
            report = 'La fecha no tiene .'
        self.ct = str(self.ct).replace(' ', '')
        self.importe = str(self.importe).replace(' ', '')

        match self.tMigracion:
            case 1:
                self.texto = self.dist + '.TRASP.CAJ.' + self.moneda + '.' + self.rec + ' A ' + self.bank + ' ' + self.fecha2
            case 2:
                match self.ETVflow:
                    case 1:
                        self.texto = self.dist + '.ENTREGA A BRINKS CIERRE ' + self.fullAsignacion + ' ' + self.fecha2
                    case 2:
                        self.texto = self.dist + '.TRASPASO ' + self.rec + ' A ' + self.bank + ' ' + self.fecha2
                    case 3:
                            self.texto = self.dist + '.DEP. DIRECTO A BANCO ' + self.fullAsignacion + ' ' + self.fecha2

                    

    def getAccountTableChildren(self, account):
        self.session.findById("wnd[0]/usr/ctxtSD_SAKNR-LOW").text = account
        self.session.findById("wnd[0]/usr/ctxtSD_BUKRS-LOW").text = "GV01"
        self.session.findById("wnd[0]/usr/ctxtSD_BUKRS-LOW").setFocus
        self.session.findById("wnd[0]/tbar[1]/btn[8]").press()

    def getRightTable(self):
        try:
            text = self.session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell/shellcont[1]").text
            writeLog('\n', 'Formato de tabla correcto', self.logPath)

        except:
            try:
                text = self.session.findById("wnd[0]/usr/lbl[0,1]").text
                writeLog('\n', 'Formato de tabla incorrecto, cambiando formato.', self.logPath)
                self.session.findById("wnd[0]/mbar/menu[5]/menu[8]").select()
            except:
                writeLog('\n', 'Formato de tabla incorrecto, no se puede cambiar.', self.logPath)

    def migrationXlsxPaste(self, asignacion, ndoc):
        rec = ''
        rec = self.rec
        rec = rec.strip()
        if 'AG.' in self.rec:
            i = rec.index('.')
            rec.insert(i, ' ')

        wbPath = currentPathParentFolder
        wbPath = os.path.join(wbPath, 'Migraciones')
        self.xlsxMigracion = self.xlsxMigracion + '.xlsx'
        wbPath = os.path.join(wbPath, self.xlsxMigracion)
        wb = load_workbook(wbPath)

        try:
            sheet = wb[self.rec]
    
            sheetList = []

            for cell in sheet['A']:
                sheetList.append(cell.value)

            for i, asig in enumerate(sheetList):
                if asig == asignacion:
                    sheet[f'K{i+1}'] = ndoc
                    wb.save(wbPath)
                    break
            
        except:
            writeLog('\n', 'No se encontró la hoja ' + self.rec + ' en el archivo de migraciones', self.logPath)

    def getBank_for_ETV(self):
        self.getFbl3nMenu()
        try:
            self.getAccountTable2()
            self.session.findById("wnd[0]/mbar/menu[5]/menu[8]").select()
            self.bank = self.session.findById("wnd[0]/usr/lbl[37,1]").text
            self.bank = self.bank.strip()
            bank_4_digits = self.bank[-4:]
            spanBankF = re.search('BANCO ', self.bank).span()[1]
            self.bank = self.bank[spanBankF:]
            splitList = re.split(r'\s', self.bank)
            bankName = splitList[0]
            self.bank = bankName + ' ' + bank_4_digits
            self.session.findById("wnd[0]/mbar/menu[5]/menu[8]").select()
            

        except Exception as e:
            print('No se pudo obtener el nombre del banco: ', e)
            self.session.EndTransaction()
            return -1


    def process(self, tMigracion, ETVflow):
        self.r = None
        self.chargeListOfNames()
        self.startSAP()
        self.chargeXlsxSheet()
        self.tMigracion = tMigracion
        self.ETVflow = ETVflow
        match self.tMigracion:
            case 1:
                self.ws2 = self.wsAg
            case 2:
                self.ws2 = self.wsDist
                
        self.xlsxRange = self.getExcelRange()
        
        print('Este es el rango del xls: ', self.xlsxRange)
        for self.r in self.xlsxRange:
            
            x =  self.subProcess_1()
            if x == -1:
                continue
            serparationMessage = f'\n\n-------------------------------- {today()} Iniciando Migracion de cuenta {self.accountNumber1} a {self.accountNumber2} --------------------------------\n\n'
            writeLog('', serparationMessage, self.logPath)
            self.subProcess_2()

    def subProcess_2_1(self):
        approvedParametersList = [[], [], [], [], [], [], [], [], [], []]
        # row = self.ws2.row(self.r)
        for row in self.ws2.iter_rows(min_row=self.r, max_row=self.r, min_col=8):
            for cell in row:
                for i, assignment in enumerate(self.approvedParametersList[7]):
                    if cell.value == assignment:
                        approvedParametersList[0].append(self.approvedParametersList[0][i])
                        approvedParametersList[1].append(self.approvedParametersList[1][i])
                        approvedParametersList[2].append(self.approvedParametersList[2][i])
                        approvedParametersList[3].append(self.approvedParametersList[3][i])
                        approvedParametersList[4].append(self.approvedParametersList[4][i])
                        approvedParametersList[5].append(self.approvedParametersList[5][i])
                        approvedParametersList[6].append(self.approvedParametersList[6][i])
                        approvedParametersList[7].append(self.approvedParametersList[7][i])
                        approvedParametersList[8].append(self.approvedParametersList[8][i])
                        approvedParametersList[9].append(self.approvedParametersList[9][i])

        self.approvedParametersList = approvedParametersList

    def subProcess_1(self):
        match self.tMigracion:
                case 1:
                    self.exec = self.ws2[f'F{self.r}'].value
                    if self.exec == 'SI':
                        pass
                    else:
                        return -1

                case 2:
                    self.exec = self.ws2[f'G{self.r}'].value
                    #self.directo = self.ws2[f'H{r}'].value
                    if self.exec == 'SI':
                        pass
                    else:
                        return -1

        if self.tMigracion == 1 or self.ETVflow == 1:
            self.accountNumber1 = self.ws2[f'C{self.r}'].value
            self.accountNumberStr1 = str(self.accountNumber1).replace(' ', '')
            self.accountNumber2 = self.ws2[f'D{self.r}'].value
            self.accountNumberStr2 = str(self.accountNumber2).replace(' ', '')
            self.bank =  self.ws2[f'E{self.r}'].value
            self.bank = str(self.bank).strip()
            

        elif self.ETVflow == 2:
            self.accountNumber1 = self.ws2[f'D{self.r}'].value
            self.accountNumberStr1 = str(self.accountNumber1).replace(' ', '')
            self.accountNumber2 = self.ws2[f'F{self.r}'].value
            self.accountNumberStr2 = str(self.accountNumber2).replace(' ', '')
            self.getBank_for_ETV()

        elif self.ETVflow == 3:
            self.accountNumber1 = self.ws2[f'C{self.r}'].value
            self.accountNumberStr1 = str(self.accountNumber1).replace(' ', '')
            self.accountNumber2 = self.ws2[f'F{self.r}'].value
            self.accountNumberStr2 = str(self.accountNumber2).replace(' ', '')
            self.getBank_for_ETV()
        
        
        self.rec =  self.ws2[f'B{self.r}'].value
        self.rec = str(self.rec)
        self.moneda = self.rec[13:16]
        self.moneda = self.moneda.replace('/', '')
        r2 = re.search('RECAUDADORA', self.rec).span()
        r2 = r2[1]
        r2+=1
        self.rec = self.rec[r2:]
        self.rec = self.rec.replace('.', '')
        self.rec = self.rec.replace('AGENCIA', '')
        self.rec = self.rec.replace('AG.', '')
        self.rec = self.rec.replace('CENTRAL', '')
        self.rec = self.rec.strip()
        
        
        
        self.txtCabDoc = 'TRASLADO A ' + self.bank

        # serparationMessage = f'\n\n-------------------------------- {today()} Iniciando Migracion de cuenta {self.accountNumber1} a {self.accountNumber2} --------------------------------\n\n'
        # writeLog('', serparationMessage, self.logPath)

        self.getFbl3nMenu()
        try:
            self.getAccountTable()
        except Exception as e:
            print('No se pudo obtener la tabla de cuentas: ', e)
            self.session.EndTransaction()
            return -1
            
        self.getRightTable()
        alert = self.session.findById('wnd[0]/sbar/pane[0]').text
        alert2 = 'No se ha seleccionado ninguna partida'
        if alert2 in alert:
            inAlert = f'No se encontró tabla de datos, revisar manualmente. CUENTA: {self.accountNumberStr1} a {self.accountNumberStr2}'
            writeLog('\n', inAlert, self.logPath)
            self.session.endTransaction()
            return -1

        match self.tMigracion:
            case 1:
                self.rec = 'AG. ' + self.rec
                parametersList = self.getWholeParametersList()
                preApprovedParametersList = self.wichMigraVerification(parametersList)
                approvedParametersList = self.wichMigraVerification2(preApprovedParametersList)
            
            case 2:
                parametersList = self.getWholeParametersList()
                approvedParametersList = self.wichMigraVerification(parametersList)
        
        self.approvedParametersList = approvedParametersList    

    
    def subProcess_2(self):

        if self.tMigracion == 2:
            self.subProcess_2_1()                        

        approvedParametersList = self.approvedParametersList
        asignacionNdocMigrated = []
        nDocsMigrated = []
        try:
            for s in range(len(approvedParametersList[0])):
                rowList = []
                rowList.append(approvedParametersList[0][s])
                rowList.append(approvedParametersList[1][s])
                rowList.append(approvedParametersList[2][s])
                rowList.append(approvedParametersList[3][s])
                rowList.append(approvedParametersList[4][s])
                rowList.append(approvedParametersList[5][s])
                rowList.append(approvedParametersList[6][s])
                rowList.append(approvedParametersList[7][s])
                rowList.append(approvedParametersList[8][s])
                rowList.append(approvedParametersList[9][s])

                self.migration(rowList)
                asignacionNdocfMigratedbyOne = []
                asignacionNdocfMigratedbyOne.append(approvedParametersList[7][s])
                asignacionNdocfMigratedbyOne.append(self.docf)
                asignacionNdocMigrated.append(asignacionNdocfMigratedbyOne)
                nDocsMigrated.append(self.docf)
                # self.migrationXlsxPaste(approvedParametersList[7][s], self.docf)
            
                # self.session.EndTransaction()
        except Exception as e:
            writeLog('\n', e, self.logPath)

        self.getFbl3nMenu()
        self.getAccountTable()
        parametersList = self.getWholeParametersList()
        self.verificationBeforeAccountChange(nDocsMigrated, approvedParametersList, parametersList)
        df = pd.DataFrame(asignacionNdocMigrated, columns = ['Asignacion', 'Ndoc'])
        ndocTOxlsx(asignacionNdocMigrated, self.rec, self.xlsxMigracion, self.logPath)
        writeLog('\n', df, self.logPath)
        serparationMessage = f'\n\n-------------------------------- Migracion de cuenta {self.accountNumber1} a {self.accountNumber2} finalizada --------------------------------\n\n'
        writeLog('', serparationMessage, self.logPath)
                           
        self.proc.kill()

  
        


       