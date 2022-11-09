from openpyxl import Workbook
from openpyxl import load_workbook
from pathlib import Path
import sys
import datetime
import os

def isNaN(num):
    return num!= num

def getCurrentPath():   
    config_name = 'myapp.cfg'
    # determine if application is a script file or frozen exe
    if getattr(sys, 'frozen', False):
        application_path = os.path.dirname(sys.executable)
    elif __file__:
        application_path = os.path.dirname(__file__)
    application_path2 = Path(application_path)
    return application_path2.parent.absolute()


currentPathParentFolder = getCurrentPath()
currentPathGrandpaFolder = Path(currentPathParentFolder).parent
logPath = os.path.join(currentPathGrandpaFolder,"log.txt")

def today():
    fullTime = datetime.datetime.now(datetime.timezone.utc) + datetime.timedelta(hours=-4)
    currenteDateStr = fullTime.strftime("%d.%m.%Y")
    return currenteDateStr

def xlsxFormatting(x):
    z1 = ''.join([today(), '-F'])
    #z = today()
    #x = f"C:\\Users\\crist\\OneDrive - UNIVERSIDAD NACIONAL DE INGENIERIA\\Venado\\Cris\\Traslado tesorería Bot\\Cuentas recaudadoras\\{z}\\CUENTAS DE CAJA IVSA.xlsx"
    wb1 = load_workbook(x)
    xlsxFormatedFolder = os.path.join(currentPathGrandpaFolder,"Cuentas recaudadoras 2")
    xlsxFormatedFolder = os.path.join(xlsxFormatedFolder,z1)
    try:        
        # a = f'C:\\Users\\crist\\OneDrive - UNIVERSIDAD NACIONAL DE INGENIERIA\\Venado\\Cris\\Traslado de tesoreria B5\\Cuentas recaudadoras 2\\{z1}'
        os.mkdir(xlsxFormatedFolder)
    except Exception as e:
        print('El archivo ya ha sido creado')
    xlsxFormatedPath = os.path.join(xlsxFormatedFolder,"CUENTAS DE CAJA IVSA.xlsx")
    # y = f"C:\\Users\\crist\\OneDrive - UNIVERSIDAD NACIONAL DE INGENIERIA\\Venado\\Cris\\Traslado de tesoreria B5\\Cuentas recaudadoras 2\\{z1}\\CUENTAS FORMATEADAS.xlsx"
    

    wb1.save(xlsxFormatedPath)
    wb2 = load_workbook(xlsxFormatedPath)
    ws2 = wb2['CAJAS RECAUDADORAS']
    mergeRangesList = []
    for i in ws2.merged_cell_ranges:
        i = str(i)
        j = i.index(':')
        #print(i, ' ', j)
        if i[0]==i[j+1] and i[0]=='D' or i[0]==i[j+1]=='E' and i[0]=='E':
            if int(i[1:j])>=3:
                mergeRangesList.append(i)
    #print(mergeRangesList[:])

    for k in mergeRangesList:
        #print(k)
        ws2.unmerge_cells(k)
        l = k.index(':')
        
        a = k[1:l]
        b = k[l+2:]
        #print(a)
        #print(b)
        a = int(a)
        b = int(b)
        a+=1
        b+=1
        
        for m in range(a, b):
            n = ''.join([k[0], str(m)])
            #print( ws2[f'{n}'])
            #print(ws2[k[:l]])
            ws2[f'{n}'] = ws2[k[:l]].value



    wb2.save(xlsxFormatedPath)
    return xlsxFormatedPath

def report(n, asignacion1, accountNumberStr2, accountNumberStr1):
    global errorList
    errorList = []

    match n:
        case 1:
            p = asignacion1 + '-' + accountNumberStr2 + '-' + accountNumberStr1 + ' fue migrado correctamente'
            errorList.append(p)
            # print(p)
            writeLog('\n', p, logPath)
        case 2:
            p = asignacion1 + '-' + accountNumberStr2 + '-' + accountNumberStr1 + ' no fue migrado correctamente, revisar manualmente'
            errorList.append(p)
            # print(p)
            writeLog('\n', p, logPath)
        case _:
            p = 'Error-ingresó un número incorrecto a la función report'
            errorList.append(p)
            # print(p)
            writeLog('\n', p, logPath)

def writeLog(s,log,rut):
    txtfolder=os.path.dirname(rut)
    pathLog=os.path.join(txtfolder, "logs.txt")
    line=s+str(log)
    print(log)
    with open(pathLog, 'a') as file:
        file.write(line)




if __name__=='__main__': 
    z = today()
    x = f"C:\\Users\\crist\\OneDrive - UNIVERSIDAD NACIONAL DE INGENIERIA\\Venado\\Cris\\Traslado tesorería Bot\\Cuentas recaudadoras\\{z}\\CUENTAS DE CAJA IVSA.xlsx" 
    xlsxFormatting(x)