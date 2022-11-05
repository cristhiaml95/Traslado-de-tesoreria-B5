import win32com.client
import subprocess
import time
from openpyxl import Workbook
from openpyxl import load_workbook
import re
from usefulFunctions import *

#poo desde video 26

class sapInterfaceJob():
    def __init__(self):
        self.k = 7
        self.i = 4
        self.j = 0

    def startSAP(self, environment):
        global sapGuiAuto, application, connection, session, paths, login
        wb = load_workbook(r'C:\Users\crist\OneDrive - UNIVERSIDAD NACIONAL DE INGENIERIA\Venado\Cris\Traslado tesorería Bot\config.xlsx')
        ws = wb['config']
        ws1 = wb['sapLogin']

        paths = {'accountPath': ws['B1'].value,
                'SAPPath': ws['B2'].value,
                'migraPath': ws['B3'].value}
               

        login = {'user': ws1['B1'].value,
                 'psw': ws1['B2'].value}
      
        proc = subprocess.Popen([paths['SAPPath'], '-new-tab'])
        time.sleep(2)
        sapGuiAuto = win32com.client.GetObject('SAPGUI')
        application = sapGuiAuto.GetScriptingEngine
        connection = application.OpenConnection(environment, True)
        session = connection.Children(0)

        session.findById("wnd[0]/usr/txtRSYST-BNAME").text = login['user']
        session.findById("wnd[0]/usr/pwdRSYST-BCODE").text = login['psw']
        session.findById("wnd[0]").sendVKey(0)

    def getFbl3nMenu(self):
        session.findById("wnd[0]/tbar[0]/okcd").text = "fbl3n"
        session.findById("wnd[0]").sendVKey(0)

    def fullProcess(self):
        global wb2, ws2, accountNumber1, accountNumberStr1, accountNumber2, accountNumberStr2, bank, asignaciones
        asignaciones = []   
        i = self.i
        k = self.k
        j = self.j
        z = today()
        x = f"C:\\Users\\crist\\OneDrive - UNIVERSIDAD NACIONAL DE INGENIERIA\\Venado\\Cris\\Traslado tesorería Bot\\Cuentas recaudadoras\\{z}\\CUENTAS DE CAJA IVSA.xlsx" 
        y = xlsxFormatting(x)
        wb2 = load_workbook(y)
        ws2 = wb2['CAJAS RECAUDADORAS']
        
        accountNumber1 = ws2[f'C{i}'].value
        accountNumberStr1 = str(accountNumber1).replace(' ', '')
        accountNumber2 = ws2[f'D{i}'].value
        accountNumberStr2 = str(accountNumber2).replace(' ', '')
        bank =  ws2[f'E{i}'].value
        bank = str(bank).strip()


        while True:
            if len(accountNumberStr1)==9 and len(accountNumberStr2)==9 and type(accountNumber1)== int and type(accountNumber2)== int:
                while True:
                    self.getFbl3nMenu()
                    self.getAccountTable()

                    if k==7:
                        self.getAssignmentInformation(k)
                        ct1 = ct
                        asignacion1 = asignacion
                        importe1 = importe
                        asignaciones = asignaciones.append(asignacion1)
                        try:
                            self.getAssignmentInformation(k+1)
                            ct2 = ct
                            asignacion2 = asignacion
                            importe2 = importe

                            match ct1:
                                case '40':
                                    importe2 = importe2.replace('-', '')
                                    if asignacion1 == asignacion2:
                                        if importe1 == importe2:
                                            report(1, asignacion1, accountNumberStr2, accountNumberStr1)                                    
                                        else:
                                            report(2, asignacion1, accountNumberStr2, accountNumberStr1)
                                        k+=1   
                                case '50':
                                    importe1 = importe1.replace('-', '')
                                    if asignacion1 == asignacion2:
                                        if importe1 == importe2:
                                            report(1, asignacion1, accountNumberStr2, accountNumberStr1)                                    
                                        else:
                                            report(2, asignacion1, accountNumberStr2, accountNumberStr1)
                                        k+=1   
                        except Exception as e:
                            print(e)
                            self.getAssignmentInformation(k)
                            self.migration()
                        


                    
                    else:
                        self.getAssignmentInformation(k-1)
                        ct0 = ct
                        asignacion0 = asignacion
                        importe0 = importe
                        try:
                            self.getAssignmentInformation(k)
                            ct1 = ct
                            asignacion1 = asignacion
                            importe1 = importe
                            asignaciones = asignaciones.append(asignacion1)       
                        except Exception as e:
                            print(e)
                            break
                        
                        self.getAssignmentInformation(k+1)
                        ct2 = ct
                        asignacion2 = asignacion
                        importe2 = importe

                        match ct1:
                                case '40':                                
                                    match ct0:
                                        case '40':
                                            match ct2:
                                                case '40':
                                                    self.getAssignmentInformation(k)
                                                    self.migration()
                                                case '50':
                                                    importe2 = importe2.replace('-', '')
                                                    if asignacion1 == asignacion2:
                                                        if importe1 == importe2:
                                                            report(1, asignacion1, accountNumberStr2, accountNumberStr1)                                    
                                                        else:
                                                            report(2, asignacion1, accountNumberStr2, accountNumberStr1)
                                                        k+=1                                                            
                                                    else:
                                                        self.getAssignmentInformation(k)
                                                        self.migration()
                                                    

                                        case '50':
                                            importe0 = importe0.replace('-', '')
                                            if asignacion1 == asignacion0:
                                                if importe1 == importe0:
                                                    report(1, asignacion1, accountNumberStr2, accountNumberStr1)
                                                else:
                                                    report(2, asignacion1, accountNumberStr2, accountNumberStr1)
                                                
                                                match ct2:
                                                    case '40':
                                                        k+=1
                                                    case '50':
                                                        if asignacion1 == asignacion2:
                                                            if importe1 == importe2:
                                                                report(1, asignacion1, accountNumberStr2, accountNumberStr1)
                                                            else:
                                                                report(2, asignacion1, accountNumberStr2, accountNumberStr1)
                                                            k+=1
                                                        else:
                                                            self.getAssignmentInformation(k)
                                                            self.migration()

                                            else:
                                                 match ct2:
                                                    case '40':
                                                        self.getAssignmentInformation(k)
                                                        self.migration()
                                                    case '50':
                                                        importe2 = importe2.replace('-', '')
                                                        if asignacion1 == asignacion2:
                                                            if importe1 == importe2:
                                                                report(1, asignacion1, accountNumberStr2, accountNumberStr1)
                                                            else:
                                                                report(2, asignacion1, accountNumberStr2, accountNumberStr1)
                                                            k+=1
                                                        else:
                                                            self.getAssignmentInformation(k)
                                                            self.migration()

                                case '50':
                                    match ct0:
                                        case '40':
                                            importe0 = importe0.replace('-', '')
                                            if asignacion1 == asignacion0:
                                                if importe1 == importe0:
                                                    report(1, asignacion1, accountNumberStr2, accountNumberStr1)
                                                else:
                                                    report(2, asignacion1, accountNumberStr2, accountNumberStr1)
                                                
                                                match ct2:
                                                    case '40':
                                                        k+=1
                                                    case '50':
                                                        k+=1
                                            else:
                                                match ct2:
                                                    case '40':
                                                        importe1 = importe1.replace('-', '')
                                                        if asignacion1 == asignacion2:
                                                            if importe1 == importe2:
                                                                report(1, asignacion1, accountNumberStr2, accountNumberStr1)
                                                            else:
                                                                report(2, asignacion1, accountNumberStr2, accountNumberStr1)
                                                            k+=1
                                                        else:
                                                            k+=1
                                                    case '50':
                                                        k+=1

                                        case '50':
                                            match ct2:
                                                case '40':
                                                    importe1 = importe1.replace('-', '')
                                                    if asignacion1 == asignacion2:
                                                        if importe1 == importe2:
                                                            report(1, asignacion1, accountNumberStr2, accountNumberStr1)
                                                        else:
                                                            report(2, asignacion1, accountNumberStr2, accountNumberStr1)
                                                        k+=1
                                                    else:
                                                        k+=1
                                                case '50':
                                                    k+=1



                i+=1

            else: 
                i+=1
                k+=1
                if k >= 2:
                    break
                else:
                    continue
# PROCESO -------------------------------------------------------------
    def migration(self):
                
                session.EndTransaction()

                session.findById("wnd[0]/tbar[0]/okcd").text = "f-02"
                session.findById("wnd[0]").sendVKey(0)

                session.findById("wnd[0]/usr/ctxtBKPF-BLDAT").text = '30.10.2022'
                session.findById("wnd[0]/usr/ctxtBKPF-BUDAT").text = '30.10.2022'
                session.findById("wnd[0]/usr/txtBKPF-XBLNR").text = rec
                session.findById("wnd[0]/usr/txtBKPF-BKTXT").text = txtCabDoc
                session.findById("wnd[0]/usr/ctxtRF05A-NEWKO").text = accountNumberStr2
                #session.findById("wnd[0]/usr/txtBKPF-MONAT").text = per
                session.findById("wnd[0]/tbar[0]/btn[0]").press()

                session.findById("wnd[0]/usr/txtBSEG-WRBTR").text = importe 
                session.findById("wnd[0]/usr/txtBSEG-ZUONR").text = asignacion
                session.findById("wnd[0]/usr/ctxtBSEG-SGTXT").text = texto
                session.findById("wnd[0]/usr/ctxtRF05A-NEWBS").text = '50'
                session.findById("wnd[0]/usr/ctxtRF05A-NEWKO").text = accountNumberStr1
                session.findById("wnd[0]/tbar[0]/btn[0]").press()

                session.findById("wnd[0]/usr/txtBSEG-WRBTR").text = importe 
                session.findById("wnd[0]/usr/txtBSEG-ZUONR").text = asignacion
                session.findById("wnd[0]/usr/ctxtBSEG-SGTXT").text = texto
                session.findById("wnd[0]/mbar/menu[0]/menu[3]").select()

                validacion = session.findById("wnd[0]/usr/txtRF05A-AZSAL").text
                validacion = str(validacion)
                validacion = validacion.replace(' ', '')
                validacion = validacion.replace('.', '')
                validacion = validacion.replace(',', '.')
                validacion = float(validacion)
                if validacion == 0:
                    print('Validación de saldo 0 correcto')
                else:
                    print(f'ERROR DE VALIDACIÓN DE SALDO 0 EN ASIGNACIÓN: {asignacion}')
                
                try:
                    session.findById("wnd[0]/tbar[0]/btn[11]").press()
                
                except Exception as e:
                    print(f'No se pudo guardar: {e}')

                session.EndTransaction()

                self.getFbl3nMenu()
                self.getAccountTable()

                
    def getAccountTable(self):
        session.findById("wnd[0]/usr/ctxtSD_SAKNR-LOW").text = accountNumberStr1
        session.findById("wnd[0]/usr/ctxtSD_BUKRS-LOW").text = "GV01"
        session.findById("wnd[0]/usr/ctxtSD_BUKRS-LOW").setFocus
        session.findById("wnd[0]/tbar[1]/btn[8]").press()

    def getAssignmentInformation(self, k):
        global rec, txtCabDoc, per, importe, asignacion, texto, a, f, c, im, fecha, ct
        
        a = f'wnd[0]/usr/lbl[9,{k}]'
        f = f'wnd[0]/usr/lbl[53,{k}]'
        c = f'wnd[0]/usr/lbl[64,{k}]'
        im = f'wnd[0]/usr/lbl[67,{k}]'
        asignacion = session.findById(a).text
        fecha = session.findById(f).text
        ct = session.findById(c).text
        importe = session.findById(im).text
        rec = session.findById('wnd[0]/usr/lbl[37,1]').text
        rec = str(rec)
        txtCabDoc = 'TRASLADO A ' + bank
        #per = 7
        print(txtCabDoc)
        r2 = re.search('RECAUDADORA', rec).span()
        r2 = r2[1]
        r2+=1
        rec = rec[r2:]
        rec = rec.strip()
        rec = rec.replace(' ', '.')
    
        asignacion = str(asignacion).replace(' ', '')
        fecha = str(fecha).replace(' ', '')
        l = fecha.index('.')
        fecha = fecha[:l+3]
        ct = str(ct).replace(' ', '')
        importe = str(importe).replace(' ', '')
        #per = str(per)

        texto = 'LP.TRASPASO ' + rec + ' A ' + bank + ' ' + fecha

if __name__=='__main__':
    environment= "QAS - EHP8 on HANA"
    bot5SapInterface = sapInterfaceJob()
    bot5SapInterface.startSAP(environment)
    bot5SapInterface.loadBankAccounts()
