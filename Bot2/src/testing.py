import win32com.client
import subprocess
import time
from openpyxl import Workbook
from openpyxl import load_workbook
import re
import os
from usefulFunctions import *

#poo desde video 26

class sapInterfaceJob():
    def __init__(self):
        self.paths = {}
        self.login = {}
    
        self.sapGuiAuto = None
        self.application = None
        self.connection = None
        self.session = None
        self.paths = None
        self.login = None
        self.proc = None
        self.rec = None
        self.txtCabDoc = None
        self.importe = None
        self.asignacion = None
        self.texto = None
        self.a = None
        self.f = None
        self.c = None
        self.im = None
        self.fecha = None
        self.ct = None
        self.check = None
        self.fechaCompleta = None
        self.wb2 = None
        self.ws2 = None
        self.accountNumber1 = None
        self.accountNumberStr1 = None
        self.accountNumber2 = None
        self.accountNumberStr2 = None
        self.bank = None
        self.last4Xlsx = None
        self.asignaciones = []
        self.ndocs = []
        self.fechas = []
        self.cts = []
        self.importes = []        
        self.textos = []
        self.checks = []
        self.fechasCompletas = []
        self.wholeParametersList = []
        self.approvedAssignments = []
        self.approvedNdocs = []
        self.approvedFechas = []
        self.approvedCts = []
        self.approvedImportes = []
        self.approvedTextos = []
        self.approvedChecks = []
        self.approvedFechasCompletas = []
        self.docf = None
        self.currentPathParentFolder = getCurrentPath()
        self.currentPathGrandpaFolder = Path(self.currentPathParentFolder).parent
        self.logPath = os.path.join(self.currentPathGrandpaFolder,"log.txt")
        self.sinceSapHunter = None
        self.xlsxFolder = None
        self.xlsxAccountList = []
        self.xlsxAccountList2 = []
        self.xlsxAccountList3 = []
        
        self.changeThePeriod = False
        self.i = 3
        self.j = 0
        self.jMax = 3
        self.k = 7
        self.rowCount = 0
        

    def startSAP(self):
        
        configXlsx=os.path.join(self.currentPathGrandpaFolder,"config.xlsx")
        wb = load_workbook(configXlsx)
        ws = wb['Rutas']
        ws1 = wb['parametrosInicio']

        self.paths = {'accountPath': ws['B1'].value,
                'SAPPath': ws['B2'].value,
                'migraPath': ws['B3'].value}
               
        
        self.login = {'user': ws1['B1'].value,
                 'psw': ws1['B2'].value,
                 'environment': ws1['B3'].value,
                 'fecha': ws1['B5'].value,
                 'periodo': ws1['B6'].value}

        if self.login['fecha'] == None:
            self.login['fecha'] = today()
        else:
            self.changeThePeriod = True
        
        self.proc = subprocess.Popen([self.paths['SAPPath'], '-new-tab'])
        time.sleep(2)
        try: 
            self.sapGuiAuto = win32com.client.GetObject('SAPGUI')
        except:
            self.proc.kill()
            self.proc = subprocess.Popen([self.paths['SAPPath'], '-new-tab'])
            self.sapGuiAuto = win32com.client.GetObject('SAPGUI')

        self.application = self.sapGuiAuto.GetScriptingEngine
        self.connection = self.application.OpenConnection(self.login['environment'], True)
        self.session = self.connection.Children(0)

        self.session.findById("wnd[0]/usr/txtRSYST-BNAME").text = self.login['user']
        self.session.findById("wnd[0]/usr/pwdRSYST-BCODE").text = self.login['psw']
        self.session.findById("wnd[0]").sendVKey(0)

    def getFbl3nMenu(self):
        self.session.EndTransaction()
        self.session.findById("wnd[0]/tbar[0]/okcd").text = "fbl3n"
        self.session.findById("wnd[0]").sendVKey(0)

    def chargeXlsxSheet(self):
        currentDay = today()
        dailyMigrationAccountsPath=os.path.join(self.currentPathGrandpaFolder,"Cuentas Recaudadoras")
        dailyMigrationAccountsPath=os.path.join(dailyMigrationAccountsPath,currentDay)
        dailyMigrationAccountsPath=os.path.join(dailyMigrationAccountsPath,"CUENTAS DE CAJA IVSA.xlsx")
        y = xlsxFormatting(dailyMigrationAccountsPath)
        self.wb2 = load_workbook(y)
        self.ws2 = self.wb2['CAJAS RECAUDADORAS']

    def getExcelRange(self):
        xlsxCellsRange = []

        while True:
            self.accountNumber1 = self.ws2[f'C{self.i}'].value
            self.accountNumberStr1 = str(self.accountNumber1).replace(' ', '')
            self.accountNumber2 = self.ws2[f'D{self.i}'].value
            self.accountNumberStr2 = str(self.accountNumber2).replace(' ', '')
            self.bank =  self.ws2[f'E{self.i}'].value
            self.bank = str(self.bank).strip()

            if len(self.accountNumberStr1)==9 and len(self.accountNumberStr2)==9 and type(self.accountNumber1)== int and type(self.accountNumber2)== int:
                xlsxCellsRange.append(self.i)
                self.i+=1
            else:
                self.i+=1
                self.j+=1
                if self.j > self.jMax:
                    break
                else:
                    continue
        return xlsxCellsRange

    def getXlsxFechaImporteList(self, bankXlsxPath):
        fechasImportes = []
        fechas = []
        importes = []
        wb = load_workbook(bankXlsxPath)
        ws = wb.active
        i = 17
        while True:
            
            fecha = ws[f'A{i}'].value
            fecha = str(fecha).replace(' ', '')
            importe = ws[f'E{i}'].value
            importe = str(importe).replace(' ', '')
            importe = importe.replace('.', '')
            importe = importe.replace(',', '.')
            try:
                importe = float(importe)
            except:
                report = 'el importe no es un número'
            if len(fecha) == 10 and type(importe) == float or type(importe) == int:
                fechas.append(fecha)
                importes.append(importe)
                i+=1
            else:
                break
        fechasImportes.append(fechas)
        fechasImportes.append(importes)

        return fechasImportes
            

    def getWholeParametersList(self):
        self.wholeParametersList = []
        self.rowCount = self.session.findById('wnd[0]/usr/cntlGRID1/shellcont/shell/shellcont[1]/shell').RowCount
        self.rowCount-=3

        for k in range(self.rowCount):
            self.k = k
            self.getRowInformation(self.k)
            self.asignaciones.append(self.asignacion)
            self.ndocs.append(self.ndoc)
            self.fechas.append(self.fecha)
            self.cts.append(self.ct)
            self.importes.append(self.importe)
            self.textos.append(self.texto)
            self.checks.append(self.check)
            self.fechasCompletas.append(self.fechaCompleta)
           
        self.wholeParametersList.append(self.asignaciones)
        self.wholeParametersList.append(self.ndocs)
        self.wholeParametersList.append(self.fechas)
        self.wholeParametersList.append(self.cts)
        self.wholeParametersList.append(self.importes)
        self.wholeParametersList.append(self.textos)
        self.wholeParametersList.append(self.checks)
        self.wholeParametersList.append(self.fechasCompletas)

        self.asignaciones = []
        self.ndocs = []
        self.cts = []
        self.importes = []
        self.textos = []
        self.checks = []
        self.fechasCompletas = []

        return self.wholeParametersList

    def wichMigraVerification(self, wholeParametersList):
        for assigment in wholeParametersList[0]:
            if wholeParametersList[0].count(assigment) == 1:               
                n = wholeParametersList[0].index(assigment)
                ndoc = wholeParametersList[1][n]
                fecha = wholeParametersList[2][n]
                ct = wholeParametersList[3][n]
                importe = wholeParametersList[4][n]
                texto = wholeParametersList[5][n]
                check = wholeParametersList[6][n]
                fechaCompleta = wholeParametersList[7][n]
                if ct == '40' and check == 0:
                    self.approvedAssignments.append(assigment)
                    self.approvedNdocs.append(ndoc)
                    self.approvedFechas.append(fecha)
                    self.approvedCts.append(ct)
                    self.approvedImportes.append(importe)
                    self.approvedTextos.append(texto)
                    self.approvedChecks.append(check)
                    self.approvedFechasCompletas.append(fechaCompleta)

        approvedParametersList = []
        approvedParametersList.append(self.approvedAssignments)
        approvedParametersList.append(self.approvedNdocs)
        approvedParametersList.append(self.approvedFechas)
        approvedParametersList.append(self.approvedCts)
        approvedParametersList.append(self.approvedImportes)
        approvedParametersList.append(self.approvedTextos)
        approvedParametersList.append(self.approvedChecks)
        approvedParametersList.append(self.approvedFechasCompletas)
        self.approvedAssignments = []
        self.approvedNdocs = []
        self.approvedFechas = []
        self.approvedCts = []
        self.approvedImportes = []
        self.approvedTextos = []
        self.approvedChecks = []
        self.approvedFechasCompletas = []
        
        return approvedParametersList

    def verificationBeforeAccountChange(self, nDocsMigrated, approvedParametersList, wholeparametersList):
        counter = 0
        for ndoc in nDocsMigrated:
            if ndoc in wholeparametersList[1]:
                n = wholeparametersList[1].index(ndoc)
                importe1 = wholeparametersList[4][n]
                importe1 = importe1.replace(' ', '')
                importe1 = importe1.replace('-', '')
                importe2 = approvedParametersList[4][counter]
                importe2 = importe2.replace(' ', '')
                importe2 = importe2.replace('-', '')
                if importe1 == importe2:
                    x = "%s La operación de asignación: %s fue migrada correctamente" %(today(), approvedParametersList[0][counter])
                    writeLog('\n', x, self.logPath)
                else:
                    y = "%s La operación de asignación: %s ERROR en importe migrado, revisar manualmente" %(today(), approvedParametersList[0][counter])
                    writeLog('\n', y, self.logPath)
            else:
                z = f'La operación de asignación: {approvedParametersList[0][counter]} FALLO en el guardado o pérdida de datos, revisar manualmente'
                y = "%s La operación de asignación: %s FALLO en el guardado o pérdida de datos, revisar manualmente" %(today(), approvedParametersList[0][counter])
                writeLog('\n', z, self.logPath)
            counter+=1

                
            
             
# PROCESO -------------------------------------------------------------
    def migration(self, rowList):                
        self.session.EndTransaction()

        self.session.findById("wnd[0]/tbar[0]/okcd").text = "f-02"
        self.session.findById("wnd[0]").sendVKey(0)

        if self.changeThePeriod:
             self.session.findById("wnd[0]/usr/ctxtBKPF-BUDAT").text = self.login['fecha']
             self.session.findById("wnd[0]/usr/txtBKPF-MONAT").text = self.login['periodo']

        self.session.findById("wnd[0]/usr/ctxtBKPF-BLDAT").text = self.login['fecha']
       
        self.session.findById("wnd[0]/usr/txtBKPF-XBLNR").text = self.rec
        self.session.findById("wnd[0]/usr/txtBKPF-BKTXT").text = self.txtCabDoc
        self.session.findById("wnd[0]/usr/ctxtRF05A-NEWKO").text = self.accountNumberStr2
        
        self.session.findById("wnd[0]/tbar[0]/btn[0]").press()
        
        try:
            self.session.findById("wnd[0]/usr/txtBSEG-WRBTR").text = rowList[4]
        except:
            periodFail = self.session.findById("wnd[0]/sbar/pane[0]").text
            self.session.endTransaction()

            raise Exception(periodFail)
        self.session.findById("wnd[0]/usr/txtBSEG-ZUONR").text = rowList[0]
        self.session.findById("wnd[0]/usr/ctxtBSEG-SGTXT").text = rowList[5]
        self.session.findById("wnd[0]/usr/ctxtRF05A-NEWBS").text = '50'
        self.session.findById("wnd[0]/usr/ctxtRF05A-NEWKO").text = self.accountNumberStr1
        self.session.findById("wnd[0]/tbar[0]/btn[0]").press()

        self.session.findById("wnd[0]/usr/txtBSEG-WRBTR").text = rowList[4]
        self.session.findById("wnd[0]/usr/txtBSEG-ZUONR").text = rowList[0]
        self.session.findById("wnd[0]/usr/ctxtBSEG-SGTXT").text = rowList[5]
        self.session.findById("wnd[0]/mbar/menu[0]/menu[3]").select()

        validacion = self.session.findById("wnd[0]/usr/txtRF05A-AZSAL").text
        validacion = str(validacion)
        validacion = validacion.replace(' ', '')
        validacion = validacion.replace('.', '')
        validacion = validacion.replace(',', '.')
        validacion = float(validacion)
        if validacion == 0:
            x = f'Validación de saldo 0 correcto en asignación: {rowList[0]}'
            writeLog('\n', x, self.logPath)
        else:
            y = f'ERROR DE VALIDACIÓN DE SALDO 0 EN ASIGNACIÓN: {rowList[0]}'
            writeLog('\n', y, self.logPath)       
   
        self.session.findById("wnd[0]/tbar[0]/btn[11]").press()

        self.docf = self.session.findById("wnd[0]/sbar/pane[0]").text
        self.docf = self.docf.replace(' ', '')
        self.docf = self.docf[4:13]
        if len(self.docf) != 9:
            self.docf = 'No hay N° doc.'

        self.session.EndTransaction()
                
    def getAccountTable(self):
        self.session.findById("wnd[0]/usr/ctxtSD_SAKNR-LOW").text = self.accountNumberStr1
        self.session.findById("wnd[0]/usr/ctxtSD_BUKRS-LOW").text = "GV01"
        self.session.findById("wnd[0]/usr/ctxtSD_BUKRS-LOW").setFocus
        self.session.findById("wnd[0]/tbar[1]/btn[8]").press()
        
        

    def getRowInformation(self, k):
        self.asignacion = None
        self.ndoc = None
        self.fecha = None
        self.ct = None
        self.importe = None
        self.txt = None
        self.check = None
        
        self.asignacion = self.session.findById('wnd[0]/usr/cntlGRID1/shellcont/shell/shellcont[1]/shell').GetCellValue(k, 'ZUONR')
        self.ndoc = self.session.findById('wnd[0]/usr/cntlGRID1/shellcont/shell/shellcont[1]/shell').GetCellValue(k, 'BELNR')
        self.fecha = self.session.findById('wnd[0]/usr/cntlGRID1/shellcont/shell/shellcont[1]/shell').GetCellValue(k, 'BLDAT')
        self.ct = self.session.findById('wnd[0]/usr/cntlGRID1/shellcont/shell/shellcont[1]/shell').GetCellValue(k, 'BSCHL')
        self.importe = self.session.findById('wnd[0]/usr/cntlGRID1/shellcont/shell/shellcont[1]/shell').GetCellValue(k, 'DMSHB')
        self.check = self.session.findById('wnd[0]/usr/cntlGRID1/shellcont/shell/shellcont[1]/shell').GetCellValue(k, 'ICO_AUGP')
        if 'Pendientes' in self.check:
            self.check = 0
        else:
            self.check = 1
        self.asignacion = str(self.asignacion).replace(' ', '')
        self.asignacion = self.asignacion[::-1]
        try:
            n = self.asignacion.index('/')
            self.asignacion = self.asignacion[n:]
        except:
            report = 'La asignación no tiene /'
        self.asignacion = self.asignacion[::-1]
        self.ndoc = str(self.ndoc).replace(' ', '')
        self.fecha = str(self.fecha).replace(' ', '')
        self.fechaCompleta = self.fecha
        try:
            l = self.fecha.index('.')
            self.fecha = self.fecha[:l+3]
        except:
            report = 'La fecha no tiene .'
        self.ct = str(self.ct).replace(' ', '')
        self.importe = str(self.importe).replace(' ', '')

        self.texto = 'LP.TRASPASO ' + 'RECAUDADORA' + ' A ' + self.bank + ' ' + self.fecha

    def getAccountTableChildren(self, account):
        self.session.findById("wnd[0]/usr/ctxtSD_SAKNR-LOW").text = account
        self.session.findById("wnd[0]/usr/ctxtSD_BUKRS-LOW").text = "GV01"
        self.session.findById("wnd[0]/usr/ctxtSD_BUKRS-LOW").setFocus
        self.session.findById("wnd[0]/tbar[1]/btn[8]").press()

    def fullProcess(self):
        self.startSAP()
        self.chargeXlsxSheet()
        xlsxRange = self.getExcelRange()
        print('Este es el rango del xlsx: ', xlsxRange)
        for r in xlsxRange:            
            self.accountNumber1 = self.ws2[f'C{r}'].value
            self.accountNumberStr1 = str(self.accountNumber1).replace(' ', '')
            self.accountNumber2 = self.ws2[f'D{r}'].value
            self.accountNumberStr2 = str(self.accountNumber2).replace(' ', '')
            self.bank =  self.ws2[f'E{r}'].value
            self.bank = str(self.bank).strip()
            self.last4Xlsx = self.bank[-4:]
            self.rec =  self.ws2[f'B{r}'].value
            self.rec = str(self.rec)
            r2 = re.search('RECAUDADORA', self.rec).span()
            r2 = r2[1]
            r2+=1
            self.rec = self.rec[r2:]
            self.rec = self.rec.strip()
            self.rec = self.rec.replace(' ', '.')
            
            self.rec = self.rec.replace('AGENCIA', 'AG')
            # self.rec = self.rec.replace('CENTRAL', 'CTL')
            self.txtCabDoc = 'TRASLADO A ' + self.bank
            self.wb2.close()

            self.getFbl3nMenu()
            try:
                self.getAccountTable()
            except Exception as e:
                print('No se pudo obtener la tabla de cuentas: ', e)
                self.session.EndTransaction()
                continue
            parametersList = self.getWholeParametersList()
            approvedParametersList = self.wichMigraVerification(parametersList)
            
            nDocsMigrated = []
            self.sinceSapHunter = 'SapHunter\plantillasSap'
            self.sinceSapHunter = os.path.join(self.sinceSapHunter,today2())
            self.xlsxFolder =  os.path.join(currentPathGrandpaFolderParent,self.sinceSapHunter)
            self.xlsxAccountList = os.listdir(self.xlsxFolder)
            for file in self.xlsxAccountList:
                if file.endswith('.xlsx'):
                    self.xlsxAccountList2.append(file)
            for f in self.xlsxAccountList2:
                self.xlsxAccountList3.append(f[:4])

            if self.last4Xlsx in self.xlsxAccountList3:
                n = self.xlsxAccountList3.index(self.last4Xlsx)
                bankXlsxPath = os.path.join(self.xlsxFolder,self.xlsxAccountList2[n])
                fechaImporteList = self.getXlsxFechaImporteList(bankXlsxPath)
                try:
                    for s in range(len(approvedParametersList[0])):
                        rowList = []
                        rowList.append(approvedParametersList[0][s])
                        rowList.append(approvedParametersList[1][s])
                        rowList.append(approvedParametersList[2][s])
                        rowList.append(approvedParametersList[3][s])
                        rowList.append(approvedParametersList[4][s])
                        rowList.append(approvedParametersList[5][s])
                        rowList.append(approvedParametersList[6][s])
                        rowList.append(approvedParametersList[7][s])

                        rowList[4] = rowList[4].replace('.', '')
                        rowList[4] = rowList[4].replace(',', '.')
                        rowList[4] = float(rowList[4])
                        

                        
                        if rowList[4] in fechaImporteList[1]:
                            m = fechaImporteList[1].index(rowList[4])
                            if rowList[7] == fechaImporteList[0][m].replace('/', '.'):
                                self.migration(rowList)                
                                nDocsMigrated.append(self.docf)
                            else:
                                log = f'La fecha no coincide con la del importe, no se hizo la migración a {rowList[0]}'
                                writeLog('\n', log, self.logPath)
                        else:
                            log = f'El importe no se encuentra en el extracto, no se hizo la migración a {rowList[0]}'
                            writeLog('\n', log, self.logPath)
                            
                        # self.migration(rowList)                
                        # nDocsMigrated.append(self.docf)
                except Exception as e:
                    writeLog('\n', e, self.logPath)

                self.getFbl3nMenu()
                self.getAccountTable()
                parametersList = self.getWholeParametersList()
                self.verificationBeforeAccountChange(nDocsMigrated, approvedParametersList, parametersList)
                #print(nDocsMigrated)
                writeLog('\n', nDocsMigrated, self.logPath)
                serparationMessage = f'\n\n------------------------------------------------------ Migracion de cuenta {self.accountNumber1} a {self.accountNumber2} finalizada -------------------------------------------------------\n\n'
                writeLog('', serparationMessage, self.logPath)
            else:
                serparationMessage = f'\n\n-------------------------------- Migracion de cuenta {self.accountNumber1} a {self.accountNumber2} ABORTADA - {self.bank} no se encuentra en plantillasSap --------------------------------\n\n'
                writeLog('', serparationMessage, self.logPath)
            self.xlsxAccountList2 = []
            self.xlsxAccountList3 = []
                           
        self.proc.kill()