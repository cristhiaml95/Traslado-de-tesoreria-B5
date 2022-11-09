import win32com.client
import subprocess
import time
from openpyxl import Workbook
from openpyxl import load_workbook
import re
import os
from usefulFunctions import *

#poo desde video 26

class sapInterfaceJob():
    def __init__(self):
        self.paths = {}
        self.login = {}
    
        self.sapGuiAuto = None
        self.application = None
        self.connection = None
        self.session = None
        self.paths = None
        self.login = None
        self.proc = None
        self.rec = None
        self.txtCabDoc = None
        self.importe = None
        self.asignacion = None
        self.texto = None
        self.a = None
        self.f = None
        self.c = None
        self.im = None
        self.fecha = None
        self.ct = None
        self.wb2 = None
        self.ws2 = None
        self.accountNumber1 = None
        self.accountNumberStr1 = None
        self.accountNumber2 = None
        self.accountNumberStr2 = None
        self.bank = None
        self.asignaciones = []
        self.ndocs = []
        self.fechas = []
        self.cts = []
        self.importes = []        
        self.textos = []
        self.wholeParametersList = []
        self.approvedAssignments = []
        self.approvedNdocs = []
        self.approvedFechas = []
        self.approvedCts = []
        self.approvedImportes = []
        self.approvedTextos = []
        self.docf = None
        self.currentPathParentFolder = getCurrentPath()
        self.currentPathGrandpaFolder = Path(self.currentPathParentFolder).parent
        self.logPath = os.path.join(self.currentPathGrandpaFolder,"log.txt")
        # self.currentPathGrandpaFolder = 'C:\\Users\\crist\\OneDrive - UNIVERSIDAD NACIONAL DE INGENIERIA\\Venado\\Cris\\Traslado de tesoreria B5'
        
        self.changeThePeriod = False
        self.i = 3
        self.j = 0
        self.jMax = 3
        self.k = 7
        

    def startSAP(self):
        
        configXlsx=os.path.join(self.currentPathGrandpaFolder,"config.xlsx")
        wb = load_workbook(configXlsx)
        ws = wb['Rutas']
        ws1 = wb['parametrosInicio']

        self.paths = {'accountPath': ws['B1'].value,
                'SAPPath': ws['B2'].value,
                'migraPath': ws['B3'].value}
               
        
        self.login = {'user': ws1['B1'].value,
                 'psw': ws1['B2'].value,
                 'environment': ws1['B3'].value,
                 'fecha': ws1['B5'].value,
                 'periodo': ws1['B6'].value}

        if self.login['fecha'] == None:
            self.login['fecha'] = today()
        else:
            self.changeThePeriod = True
        

        
        

      
        self.proc = subprocess.Popen([self.paths['SAPPath'], '-new-tab'])
        time.sleep(2)
        try: 
            self.sapGuiAuto = win32com.client.GetObject('SAPGUI')
        except:
            self.proc.kill()
            self.proc = subprocess.Popen([self.paths['SAPPath'], '-new-tab'])
            self.sapGuiAuto = win32com.client.GetObject('SAPGUI')

        self.application = self.sapGuiAuto.GetScriptingEngine
        self.connection = self.application.OpenConnection(self.login['environment'], True)
        self.session = self.connection.Children(0)

        self.session.findById("wnd[0]/usr/txtRSYST-BNAME").text = self.login['user']
        self.session.findById("wnd[0]/usr/pwdRSYST-BCODE").text = self.login['psw']
        self.session.findById("wnd[0]").sendVKey(0)

    def getFbl3nMenu(self):
        self.session.EndTransaction()
        self.session.findById("wnd[0]/tbar[0]/okcd").text = "fbl3n"
        self.session.findById("wnd[0]").sendVKey(0)

    def chargeXlsxSheet(self):
        currentDay = today()
        dailyMigrationAccountsPath=os.path.join(self.currentPathGrandpaFolder,"Cuentas Recaudadoras")
        dailyMigrationAccountsPath=os.path.join(dailyMigrationAccountsPath,currentDay)
        dailyMigrationAccountsPath=os.path.join(dailyMigrationAccountsPath,"CUENTAS DE CAJA IVSA.xlsx")
        # x = f"C:\\Users\\crist\\OneDrive - UNIVERSIDAD NACIONAL DE INGENIERIA\\Venado\\Cris\\Traslado de tesoreria B5\\Cuentas recaudadoras\\{currentDay}\\CUENTAS DE CAJA IVSA.xlsx" 
        y = xlsxFormatting(dailyMigrationAccountsPath)
        self.wb2 = load_workbook(y)
        self.ws2 = self.wb2['CAJAS RECAUDADORAS']

    def getExcelRange(self):
      
        # z = today()
        # x = f"C:\\Users\\crist\\OneDrive - UNIVERSIDAD NACIONAL DE INGENIERIA\\Venado\\Cris\\Traslado de tesoreria B5\\Cuentas recaudadoras\\{z}\\CUENTAS DE CAJA IVSA.xlsx" 
        # y = xlsxFormatting(x)
        # self.wb2 = load_workbook(y)
        # self.ws2 = self.wb2['CAJAS RECAUDADORAS']
        xlsxCellsRange = []

        while True:
            self.accountNumber1 = self.ws2[f'C{self.i}'].value
            self.accountNumberStr1 = str(self.accountNumber1).replace(' ', '')
            self.accountNumber2 = self.ws2[f'D{self.i}'].value
            self.accountNumberStr2 = str(self.accountNumber2).replace(' ', '')
            self.bank =  self.ws2[f'E{self.i}'].value
            self.bank = str(self.bank).strip()

            if len(self.accountNumberStr1)==9 and len(self.accountNumberStr2)==9 and type(self.accountNumber1)== int and type(self.accountNumber2)== int:
                xlsxCellsRange.append(self.i)
                self.i+=1
            else:
                self.i+=1
                self.j+=1
                if self.j > self.jMax:
                    break
                else:
                    continue
        return xlsxCellsRange
            

    def getWholeParametersList(self):
        self.wholeParametersList = []
        while True:
            try:
                self.getRowInformation(self.k)
                self.asignaciones.append(self.asignacion)
                self.ndocs.append(self.ndoc)
                self.fechas.append(self.fecha)
                self.cts.append(self.ct)
                self.importes.append(self.importe)
                self.textos.append(self.texto)
                if self.k == 44:
                    self.session.findById("wnd[0]/usr").verticalScrollbar.position = '26'
                    self.k = 18
                self.k+=1
            except Exception as e:
                # print(e)
                #print('Tabla migrada completa')
                if self.k == 7:
                    writeLog('\n', 'No hay filas en la table', self.logPath)
                else:                    
                    writeLog('\n','Tabla leida correctamente', self.logPath)
                break
        self.wholeParametersList.append(self.asignaciones)
        self.wholeParametersList.append(self.ndocs)
        self.wholeParametersList.append(self.fechas)
        self.wholeParametersList.append(self.cts)
        self.wholeParametersList.append(self.importes)
        self.wholeParametersList.append(self.textos)

        self.k = 7
        self.asignaciones = []
        self.ndocs = []
        self.cts = []
        self.importes = []
        #writeLog('\n', self.wholeParametersList[0], self.currentPathGrandpaFolder)

        return self.wholeParametersList

    def wichMigraVerification(self, wholeParametersList):
        for assigment in wholeParametersList[0]:
            if wholeParametersList[0].count(assigment) == 1:               
                n = wholeParametersList[0].index(assigment)
                ndoc = wholeParametersList[1][n]
                fecha = wholeParametersList[2][n]
                ct = wholeParametersList[3][n]
                importe = wholeParametersList[4][n]
                texto = wholeParametersList[5][n]

                self.approvedAssignments.append(assigment)
                self.approvedNdocs.append(ndoc)
                self.approvedFechas.append(fecha)
                self.approvedCts.append(ct)
                self.approvedImportes.append(importe)
                self.approvedTextos.append(texto)

        approvedParametersList = []
        approvedParametersList.append(self.approvedAssignments)
        approvedParametersList.append(self.approvedNdocs)
        approvedParametersList.append(self.approvedFechas)
        approvedParametersList.append(self.approvedCts)
        approvedParametersList.append(self.approvedImportes)
        approvedParametersList.append(self.approvedTextos)

        return approvedParametersList

    def verificationBeforeAccountChange(self, nDocsMigrated, approvedParametersList, wholeparametersList):
        counter = 0
        for ndoc in nDocsMigrated:
            if ndoc in wholeparametersList[1]:
                n = wholeparametersList[1].index(ndoc)
                importe1 = wholeparametersList[4][n]
                importe1 = importe1.replace(' ', '')
                importe1 = importe1.replace('-', '')
                importe2 = approvedParametersList[4][counter]
                importe2 = importe2.replace(' ', '')
                importe2 = importe2.replace('-', '')
                if importe1 == importe2:
                    x = "%s La operación de asignación: %s fue migrada correctamente" %(today(), approvedParametersList[0][counter])
                    writeLog('\n', x, self.logPath)
                else:
                    y = "%s La operación de asignación: %s ERROR en importe migrado, revisar manualmente" %(today(), approvedParametersList[0][counter])
                    writeLog('\n', y, self.logPath)
            else:
                z = f'La operación de asignación: {approvedParametersList[0][counter]} FALLO en el guardado o pérdida de datos, revisar manualmente'
                y = "%s La operación de asignación: %s FALLO en el guardado o pérdida de datos, revisar manualmente" %(today(), approvedParametersList[0][counter])
                writeLog('\n', z, self.logPath)
            counter+=1

                
            
             
# PROCESO -------------------------------------------------------------
    def migration(self, rowList):                
        self.session.EndTransaction()

        self.session.findById("wnd[0]/tbar[0]/okcd").text = "f-02"
        self.session.findById("wnd[0]").sendVKey(0)

        if self.changeThePeriod:
             self.session.findById("wnd[0]/usr/ctxtBKPF-BUDAT").text = self.login['fecha']
             self.session.findById("wnd[0]/usr/txtBKPF-MONAT").text = self.login['periodo']

        self.session.findById("wnd[0]/usr/ctxtBKPF-BLDAT").text = self.login['fecha']
       
        self.session.findById("wnd[0]/usr/txtBKPF-XBLNR").text = self.rec
        self.session.findById("wnd[0]/usr/txtBKPF-BKTXT").text = self.txtCabDoc
        self.session.findById("wnd[0]/usr/ctxtRF05A-NEWKO").text = self.accountNumberStr2
        
        self.session.findById("wnd[0]/tbar[0]/btn[0]").press()
        
        try:
            self.session.findById("wnd[0]/usr/txtBSEG-WRBTR").text = rowList[4]
        except:
            periodFail = self.session.findById("wnd[0]/sbar/pane[0]").text
            self.session.endTransaction()
            # self.session.endTransaction()
            # self.session.findById("wnd[0]/usr/btnSTARTBUTTON").press()

            raise Exception(periodFail)
        self.session.findById("wnd[0]/usr/txtBSEG-ZUONR").text = rowList[0]
        self.session.findById("wnd[0]/usr/ctxtBSEG-SGTXT").text = rowList[5]
        self.session.findById("wnd[0]/usr/ctxtRF05A-NEWBS").text = '50'
        self.session.findById("wnd[0]/usr/ctxtRF05A-NEWKO").text = self.accountNumberStr1
        self.session.findById("wnd[0]/tbar[0]/btn[0]").press()

        self.session.findById("wnd[0]/usr/txtBSEG-WRBTR").text = rowList[4]
        self.session.findById("wnd[0]/usr/txtBSEG-ZUONR").text = rowList[0]
        self.session.findById("wnd[0]/usr/ctxtBSEG-SGTXT").text = rowList[5]
        self.session.findById("wnd[0]/mbar/menu[0]/menu[3]").select()

        validacion = self.session.findById("wnd[0]/usr/txtRF05A-AZSAL").text
        validacion = str(validacion)
        validacion = validacion.replace(' ', '')
        validacion = validacion.replace('.', '')
        validacion = validacion.replace(',', '.')
        validacion = float(validacion)
        if validacion == 0:
            x = 'Validación de saldo 0 correcto'
            writeLog('\n', x, self.logPath)
        else:
            y = f'ERROR DE VALIDACIÓN DE SALDO 0 EN ASIGNACIÓN: {self.asignacion}'
            writeLog('\n', y, self.logPath)
        
        try:
            self.session.findById("wnd[0]/tbar[0]/btn[11]").press()
        #--------------este try está por las huevas-------------------
        except Exception as e:
            z = f'No se pudo guardar: {e}'
            writeLog('\n', z, self.logPath)

        self.docf = self.session.findById("wnd[0]/sbar/pane[0]").text
        self.docf = self.docf.replace(' ', '')
        self.docf = self.docf[4:13]
        if len(self.docf) != 9:
            self.docf = 'No hay N° doc.'
        
        #writeLog('\n', self.docf, self.currentPathGrandpaFolder)

        self.session.EndTransaction()
                
    def getAccountTable(self):
        self.session.findById("wnd[0]/usr/ctxtSD_SAKNR-LOW").text = self.accountNumberStr1
        self.session.findById("wnd[0]/usr/ctxtSD_BUKRS-LOW").text = "GV01"
        self.session.findById("wnd[0]/usr/ctxtSD_BUKRS-LOW").setFocus
        self.session.findById("wnd[0]/tbar[1]/btn[8]").press()
        self.rec = self.session.findById('wnd[0]/usr/lbl[37,1]').text
        self.rec = str(self.rec)
        r2 = re.search('RECAUDADORA', self.rec).span()
        r2 = r2[1]
        r2+=1
        self.rec = self.rec[r2:]
        self.rec = self.rec.strip()
        self.rec = self.rec.replace(' ', '.')
        
        self.rec = self.rec.replace('AGENCIA', 'AG')
        self.rec = self.rec.replace('CENTRAL', 'CTL')
        self.txtCabDoc = 'TRASLADO A ' + self.bank

    def getRowInformation(self, k):
        
        self.a = f'wnd[0]/usr/lbl[9,{k}]'
        self.nd = f'wnd[0]/usr/lbl[28,{k}]'
        self.f = f'wnd[0]/usr/lbl[53,{k}]'
        self.c = f'wnd[0]/usr/lbl[64,{k}]'
        self.im = f'wnd[0]/usr/lbl[67,{k}]'
        self.asignacion = self.session.findById(self.a).text
        self.ndoc = self.session.findById(self.nd).text
        self.fecha = self.session.findById(self.f).text
        self.ct = self.session.findById(self.c).text
        self.importe = self.session.findById(self.im).text
        # self.rec = self.session.findById('wnd[0]/usr/lbl[37,1]').text
        # self.rec = str(self.rec)
        # self.txtCabDoc = 'TRASLADO A ' + self.bank
        # print(self.txtCabDoc)
        # r2 = re.search('RECAUDADORA', self.rec).span()
        # r2 = r2[1]
        # r2+=1
        # self.rec = self.rec[r2:]
        # self.rec = self.rec.strip()
        # self.rec = self.rec.replace(' ', '.')
    
        self.asignacion = str(self.asignacion).replace(' ', '')
        self.asignacion = self.asignacion[::-1]
        n = self.asignacion.index('/')
        self.asignacion = self.asignacion[n:]
        self.asignacion = self.asignacion[::-1]
        self.ndoc = str(self.ndoc).replace(' ', '')
        self.fecha = str(self.fecha).replace(' ', '')
        l = self.fecha.index('.')
        self.fecha = self.fecha[:l+3]
        self.ct = str(self.ct).replace(' ', '')
        self.importe = str(self.importe).replace(' ', '')
        #per = str(per)

        self.texto = 'LP.TRASPASO ' + self.rec + ' A ' + self.bank + ' ' + self.fecha

    def fullProcess(self):
        # environment = "QAS - EHP8 on HANA"
        self.startSAP()
        self.chargeXlsxSheet()
        xlsxRange = self.getExcelRange()
        print('Este es el rango del xlsx: ', xlsxRange)
        for r in xlsxRange:            
            self.accountNumber1 = self.ws2[f'C{r}'].value
            self.accountNumberStr1 = str(self.accountNumber1).replace(' ', '')
            self.accountNumber2 = self.ws2[f'D{r}'].value
            self.accountNumberStr2 = str(self.accountNumber2).replace(' ', '')
            self.bank =  self.ws2[f'E{r}'].value
            self.bank = str(self.bank).strip()

            self.getFbl3nMenu()
            self.getAccountTable()
            parametersList = self.getWholeParametersList()
            approvedParametersList = self.wichMigraVerification(parametersList)
            x = 'Lista de aprobados para la migración: ' + str(approvedParametersList)
            #writeLog('\n', x, self.currentPathGrandpaFolder)
            #print(approvedParametersList)
            nDocsMigrated = []
            try:
                for s in range(len(approvedParametersList[0])):
                    rowList = []
                    rowList.append(approvedParametersList[0][s])
                    rowList.append(approvedParametersList[1][s])
                    rowList.append(approvedParametersList[2][s])
                    rowList.append(approvedParametersList[3][s])
                    rowList.append(approvedParametersList[4][s])
                    rowList.append(approvedParametersList[5][s])

                    self.migration(rowList)                
                    nDocsMigrated.append(self.docf)
                    # self.session.EndTransaction()
            except Exception as e:
                writeLog('\n', e, self.logPath)

            self.getFbl3nMenu()
            self.getAccountTable()
            parametersList = self.getWholeParametersList()
            self.verificationBeforeAccountChange(nDocsMigrated, approvedParametersList, parametersList)
            #print(nDocsMigrated)
            writeLog('\n', nDocsMigrated, self.logPath)
            serparationMessage = f'\n\n-------------------------------- Migracion de cuenta {self.accountNumber1} a {self.accountNumber2} finalizada --------------------------------\n\n'
            writeLog('', serparationMessage, self.logPath)
                           
        self.proc.kill()
